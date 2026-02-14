"""产品导入导出服务。

业务范围：
1. 导出产品库 Excel 模板：生成表头、示例说明区、数据区和“说明”页。
2. 导入产品库 Excel：识别模板表头、逐行校验、收集错误/警告，并按名称执行新增或更新。

数据约束：
- 产品名称在单次导入文件内必须唯一。
- 单价波动允许输入小数或百分数（0-100 会自动换算为 0-1 小数）。
"""

from datetime import datetime
import re
from decimal import Decimal, InvalidOperation
from typing import Any
import io

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter

from app.db.serializers import encode_for_mongo
from app.services.unit_rules import normalize_unit_input, quantity_step_for_unit


PRODUCT_TEMPLATE_HEADERS = [
    "name",
    "category_name",
    "unit",
    "base_price",
    "volatility",
    "item_quantity_range_min",
    "item_quantity_range_max",
    "is_active",
]
"""导入导出模板的标准字段顺序（英文键）。"""

PRODUCT_HEADER_LABELS = {
    "name": "产品名称",
    "category_name": "品类名称",
    "unit": "单位",
    "base_price": "单价(元)",
    "volatility": "单价波动(%)",
    "item_quantity_range_min": "采购数量范围-最小",
    "item_quantity_range_max": "采购数量范围-最大",
    "is_active": "启用状态",
}
"""标准字段与中文表头的映射。"""

PRODUCT_HEADER_ALIASES = {
    **{label: key for key, label in PRODUCT_HEADER_LABELS.items()},
    **{key: key for key in PRODUCT_TEMPLATE_HEADERS},
}
"""表头兼容映射：支持中文表头和英文键两种形式。"""

PRODUCT_FIELD_NOTES = [
    {"label": "单价(元)", "desc": "基础单价固定按两位小数处理，且最小值为 0.01。"},
    {"label": "单价波动(%)", "desc": "控制单价浮动比例，输入 0-100 会自动换算为 0-1。"},
    {"label": "采购数量范围-最小", "desc": "单次采购数量下限，生成计划时不会低于该值。"},
    {"label": "采购数量范围-最大", "desc": "单次采购数量上限，生成计划时不会高于该值。"},
    {"label": "单位", "desc": "后端会自动标准化单位；可分割单位(如克/斤/千克/毫升)数量按 0.1 步进，其它按 1。"},
]
"""导出模板“说明”页中展示的重点字段说明。"""


def normalize_base_price(value: Decimal) -> Decimal:
    """规范化基础单价：两位小数，最小值 0.01。"""
    rounded = value.quantize(Decimal("0.01"))
    return Decimal("0.01") if rounded < Decimal("0.01") else rounded


def _normalize_cell(value: Any) -> Any:
    """标准化单元格值：字符串去首尾空白，其它类型原样返回。"""
    if isinstance(value, str):
        return value.strip()
    return value


def _normalize_header_text(value: Any) -> str:
    """标准化表头文本，兼容全角/空白/标点差异。"""
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    replacements = {
        "（": "(",
        "）": ")",
        "％": "%",
        "—": "-",
        "–": "-",
        "－": "-",
        "−": "-",
        "：": ":",
        "\u00A0": " ",
        "\u200B": "",
        "\u200C": "",
        "\u200D": "",
        "\uFEFF": "",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _build_normalized_header_aliases() -> dict[str, str]:
    """构建标准化后的表头别名映射，便于匹配。"""
    mapping: dict[str, str] = {}
    for raw, key in PRODUCT_HEADER_ALIASES.items():
        normalized = _normalize_header_text(raw)
        if normalized:
            mapping[normalized] = key
    return mapping


NORMALIZED_HEADER_ALIASES = _build_normalized_header_aliases()


def _parse_decimal(value: Any, row_idx: int, field: str, errors: list[dict[str, Any]]) -> Decimal | None:
    """解析 Decimal；为空或格式非法时记录错误并返回 None。"""
    if value is None or str(value).strip() == "":
        errors.append({"row": row_idx, "field": field, "message": "不能为空", "value": value})
        return None
    try:
        return Decimal(str(value))
    except InvalidOperation:
        errors.append({"row": row_idx, "field": field, "message": "格式无效", "value": value})
        return None


def _is_multiple_of_step(value: Decimal, step: Decimal) -> bool:
    """判断值是否为步进值的整数倍。"""
    if step <= 0:
        return True
    remainder = (value / step) % 1
    return remainder == 0


def _parse_bool(value: Any) -> bool | None:
    """兼容中英文/数字布尔输入，无法识别时返回 None。"""
    if isinstance(value, bool):
        return value
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if value == 1:
            return True
        if value == 0:
            return False
    if isinstance(value, str):
        text = value.strip().lower()
        if text in {"true", "t", "yes", "y", "1", "启用", "是"}:
            return True
        if text in {"false", "f", "no", "n", "0", "停用", "禁用", "否", "已作废"}:
            return False
    return None


def build_products_workbook(products: list[dict[str, Any]]) -> Workbook:
    """生成产品库导出工作簿。

    - 主表“产品库”：表头 + 数据区 + 右侧说明区。
    - 子表“说明”：导入规则与重点字段解释。
    """
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = "产品库"
    ws.append([PRODUCT_HEADER_LABELS.get(field, field) for field in PRODUCT_TEMPLATE_HEADERS])
    ws.freeze_panes = "A2"

    header_columns = len(PRODUCT_TEMPLATE_HEADERS)
    wrap_left = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    header_align = Alignment(horizontal="center", vertical="center")
    body_align = Alignment(horizontal="center", vertical="center")
    note_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    bold_font = Font(bold=True)

    notes = [
        ("说明", "表头不可修改。"),
        ("产品名称", "名称唯一，按名称更新或新增。"),
        ("单价波动(%)", "控制单价允许的浮动比例，百分数(0-100)。"),
        ("采购数量范围-最小", "单次采购下限。"),
        ("采购数量范围-最大", "单次采购上限。"),
        ("单位", "单位会自动标准化；克/斤/千克/毫升/升等按 0.1 步进，其他按 1。"),
        ("启用状态", "启用=当前可用；已作废=停止使用，仅用于历史数据。"),
    ]
    note_label_col = header_columns + 2
    note_desc_col = header_columns + 3
    for idx, (label, desc) in enumerate(notes, start=1):
        label_cell = ws.cell(row=idx, column=note_label_col, value=label)
        desc_cell = ws.cell(row=idx, column=note_desc_col, value=desc)
        label_cell.alignment = wrap_left
        desc_cell.alignment = wrap_left
        label_cell.fill = note_fill
        desc_cell.fill = note_fill
        label_cell.font = bold_font

    data_start_row = 2
    for idx, doc in enumerate(products):
        row_idx = data_start_row + idx
        item_range = doc.get("item_quantity_range") or {}
        raw_unit = doc.get("unit", "")
        if raw_unit:
            try:
                display_unit = normalize_unit_input(raw_unit)
            except ValueError:
                display_unit = str(raw_unit).strip()
        else:
            display_unit = ""

        def _num(value: Any):
            """将数值转为适合 Excel 展示的 int/float，空值保持为空字符串。"""
            if value is None or value == "":
                return ""
            num = Decimal(str(value))
            if num == num.to_integral():
                return int(num)
            return float(num)

        row_values = [
            doc.get("name", ""),
            doc.get("category_name", "") or doc.get("category", ""),
            display_unit,
            _num(doc.get("base_price")),
            _num(Decimal(str(doc.get("volatility"))) * Decimal("100"))
            if doc.get("volatility") is not None
            else "",
            _num(item_range.get("min", "")),
            _num(item_range.get("max", "")),
            "启用" if not doc.get("is_deleted", False) else "已作废",
        ]
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

        def _apply_number(cell, value, decimals_format):
            """将单元格值统一为数值并设置显示格式（不强制补零）。"""
            if value == "" or value is None:
                return
            num = Decimal(str(value))
            cell.value = float(num)
            cell.number_format = decimals_format

        _apply_number(ws.cell(row=row_idx, column=4), ws.cell(row=row_idx, column=4).value, "0.00")
        _apply_number(ws.cell(row=row_idx, column=5), ws.cell(row=row_idx, column=5).value, "0.##")
        _apply_number(ws.cell(row=row_idx, column=6), ws.cell(row=row_idx, column=6).value, "0.##")
        _apply_number(ws.cell(row=row_idx, column=7), ws.cell(row=row_idx, column=7).value, "0.##")

    header_widths = {
        "name": 26,
        "category_name": 16,
        "unit": 10,
        "base_price": 12,
        "volatility": 12,
        "item_quantity_range_min": 18,
        "item_quantity_range_max": 18,
        "is_active": 10,
    }
    for idx, header in enumerate(PRODUCT_TEMPLATE_HEADERS, start=1):
        width = header_widths.get(header, max(12, len(header) + 2))
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.column_dimensions[get_column_letter(note_label_col)].width = 22
    ws.column_dimensions[get_column_letter(note_desc_col)].width = 80

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    table_last_row = data_start_row + max(len(products), 1) - 1
    for row in ws.iter_rows(min_row=1, max_row=table_last_row, min_col=1, max_col=header_columns):
        for cell in row:
            cell.border = border
            if cell.row == 1:
                cell.alignment = header_align
                cell.font = bold_font
            else:
                cell.alignment = body_align

    for row in ws.iter_rows(min_row=1, max_row=len(notes), min_col=note_label_col, max_col=note_desc_col):
        for cell in row:
            cell.border = border

    info = wb.create_sheet("说明")
    info.append(["使用说明"])
    info.append(["1. 使用导出的模板填写后导入，表头不可修改。"])
    info.append(["2. 以产品名称作为唯一标识，存在则更新，不存在则新增。"])
    info.append([])
    info.append(["字段说明（仅列出需重点关注字段）"])
    for item in PRODUCT_FIELD_NOTES:
        info.append([item["label"], "：", item["desc"]])
    info.column_dimensions["A"].width = 22
    info.column_dimensions["B"].width = 4
    info.column_dimensions["C"].width = 80
    for row in info.iter_rows(min_row=1, max_row=info.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.border = border
            cell.alignment = header_align if cell.row == 1 else wrap_left
        if row[0].row > 1:
            row[0].font = bold_font
    info.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    return wb


async def import_products_from_xlsx(
    db: Any,
    payload: bytes,
    dry_run: bool = False,
) -> dict[str, Any]:
    """导入产品库 Excel。

    处理流程：
    1. 自动定位并校验模板表头。
    2. 逐行校验字段，收集 errors/warnings。
    3. 无 errors 时按产品名称执行 upsert（dry_run 时仅统计不落库）。
    """
    wb = load_workbook(io.BytesIO(payload), data_only=True)
    ws = wb["产品库"] if "产品库" in wb.sheetnames else wb.active

    header_row_idx = 0
    header_columns = len(PRODUCT_TEMPLATE_HEADERS)
    header_map: dict[str, int] = {}
    header_debug: list[dict[str, Any]] = []
    max_scan_rows = min(ws.max_row, 50)
    max_scan_cols = min(ws.max_column or header_columns, header_columns + 12)
    for idx in range(1, max_scan_rows + 1):
        row_map: dict[str, int] = {}
        row_texts: list[str] = []
        for col in range(1, max_scan_cols + 1):
            text = _normalize_header_text(ws.cell(row=idx, column=col).value)
            row_texts.append(text)
            if not text:
                continue
            header_key = NORMALIZED_HEADER_ALIASES.get(text, "")
            if header_key in PRODUCT_TEMPLATE_HEADERS and header_key not in row_map:
                row_map[header_key] = col
        if any(row_texts):
            header_debug.append({"row": idx, "texts": row_texts, "matched": list(row_map.keys())})
        if all(key in row_map for key in PRODUCT_TEMPLATE_HEADERS):
            header_row_idx = idx
            header_map = row_map
            break
    if header_row_idx == 0:
        raise HTTPException(
            status_code=400,
            detail={
                "message": "模板表头不匹配，请使用最新模板",
                "scanned": header_debug[:5],
            },
        )

    errors: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    raw_rows: list[dict[str, Any]] = []
    seen_names: set[str] = set()
    invalid_rows: set[int] = set()

    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        row_cells = [
            ws.cell(row=row_idx, column=header_map[header]).value
            for header in PRODUCT_TEMPLATE_HEADERS
        ]
        if all(cell is None or str(cell).strip() == "" for cell in row_cells):
            continue
        row_data = {header: _normalize_cell(value) for header, value in zip(PRODUCT_TEMPLATE_HEADERS, row_cells)}
        name = row_data.get("name") or ""
        if not str(name).strip():
            errors.append({"row": row_idx, "field": "name", "message": "不能为空", "value": name})
            invalid_rows.add(row_idx)
        else:
            trimmed = str(name).strip()
            if trimmed in seen_names:
                errors.append({"row": row_idx, "field": "name", "message": "名称重复", "value": name})
                invalid_rows.add(row_idx)
            seen_names.add(trimmed)
            row_data["name"] = trimmed
        raw_rows.append({"row": row_idx, "data": row_data})

    category_names = {row["data"].get("category_name") for row in raw_rows if row["data"].get("category_name")}
    category_map: dict[str, dict[str, Any]] = {}
    if category_names:
        cursor = db["categories"].find({"name": {"$in": list(category_names)}})
        async for doc in cursor:
            category_map[doc.get("name", "")] = doc

    prepared: list[dict[str, Any]] = []
    for row in raw_rows:
        row_idx = row["row"]
        data = row["data"]
        if row_idx in invalid_rows:
            continue
        row_error_count = len(errors)

        category_name = data.get("category_name")
        if not category_name:
            errors.append({"row": row_idx, "field": "category_name", "message": "不能为空", "value": category_name})
        else:
            category_doc = category_map.get(str(category_name))
            if not category_doc:
                errors.append({"row": row_idx, "field": "category_name", "message": "品类不存在", "value": category_name})
            elif not category_doc.get("is_active", True):
                errors.append({"row": row_idx, "field": "category_name", "message": "品类已停用", "value": category_name})

        unit = data.get("unit")
        if unit is None or not str(unit).strip():
            errors.append({"row": row_idx, "field": "unit", "message": "不能为空", "value": unit})
        else:
            try:
                data["unit"] = normalize_unit_input(str(unit))
            except ValueError as exc:
                errors.append({"row": row_idx, "field": "unit", "message": str(exc), "value": unit})

        base_price = _parse_decimal(data.get("base_price"), row_idx, "base_price", errors)
        volatility = _parse_decimal(data.get("volatility"), row_idx, "volatility", errors)
        min_value = _parse_decimal(data.get("item_quantity_range_min"), row_idx, "item_quantity_range_min", errors)
        max_value = _parse_decimal(data.get("item_quantity_range_max"), row_idx, "item_quantity_range_max", errors)

        is_active_raw = data.get("is_active")
        is_active = _parse_bool(is_active_raw)
        if is_active is None:
            errors.append({"row": row_idx, "field": "is_active", "message": "值无效", "value": is_active_raw})

        if base_price is not None and base_price < Decimal("0.01"):
            errors.append({"row": row_idx, "field": "base_price", "message": "必须大于等于 0.01", "value": base_price})
        if base_price is not None:
            normalized_price = normalize_base_price(base_price)
            if normalized_price != base_price:
                warnings.append({"row": row_idx, "field": "base_price", "message": "已按两位小数自动修正", "value": base_price})
            base_price = normalized_price
        if volatility is not None:
            if volatility < 0:
                errors.append({"row": row_idx, "field": "volatility", "message": "必须大于等于 0", "value": volatility})
            elif volatility > 1:
                if volatility <= 100:
                    volatility = (volatility / Decimal("100")).quantize(Decimal("0.0001"))
                else:
                    errors.append({"row": row_idx, "field": "volatility", "message": "超过 100%", "value": volatility})
        if min_value is not None and max_value is not None and min_value > max_value:
            errors.append({"row": row_idx, "field": "item_quantity_range", "message": "最小值不能大于最大值", "value": f"{min_value}-{max_value}"})
        quantity_step = quantity_step_for_unit(data.get("unit", ""))
        if min_value is not None and not _is_multiple_of_step(min_value, quantity_step):
            warnings.append({"row": row_idx, "field": "item_quantity_range_min", "message": "最小值不是步进的整数倍", "value": min_value})
        if max_value is not None and not _is_multiple_of_step(max_value, quantity_step):
            warnings.append({"row": row_idx, "field": "item_quantity_range_max", "message": "最大值不是步进的整数倍", "value": max_value})

        if len(errors) > row_error_count:
            continue

        category_doc = category_map.get(str(category_name))
        prepared.append(
            {
                "row": row_idx,
                "name": data.get("name"),
                "category_id": str(category_doc["_id"]) if category_doc else "",
                "category_name": category_doc.get("name", "") if category_doc else "",
                "unit": str(unit).strip(),
                "base_price": base_price,
                "volatility": volatility,
                "item_quantity_range": {"min": min_value, "max": max_value},
                "is_deleted": not bool(is_active),
            }
        )

    imported_names = [row["data"].get("name") for row in raw_rows if row["data"].get("name")]
    imported_name_set = {str(name).strip() for name in imported_names if str(name).strip()}
    deactivate_candidates: list[str] = []
    cursor = db["products"].find({"is_deleted": False}, {"name": 1})
    async for doc in cursor:
        name = str(doc.get("name", "")).strip()
        if name and name not in imported_name_set:
            deactivate_candidates.append(name)

    if errors:
        skipped = max(len(raw_rows) - len(prepared), 0)
        return {
            "total": len(raw_rows),
            "valid": len(prepared),
            "skipped": skipped,
            "created": 0,
            "updated": 0,
            "warnings": warnings,
            "errors": errors,
            "deactivate_candidates": deactivate_candidates,
            "applied": False,
            "dry_run": dry_run,
        }

    names = [row["name"] for row in prepared]
    existing_map: dict[str, dict[str, Any]] = {}
    if names:
        cursor = db["products"].find({"name": {"$in": names}})
        async for doc in cursor:
            existing_map[doc.get("name", "")] = doc

    created = 0
    updated = 0
    deactivated = 0
    if not dry_run:
        now = datetime.utcnow()
        for row in prepared:
            existing = existing_map.get(row["name"])
            if existing:
                update_doc = {
                    "category_id": row["category_id"],
                    "category_name": row["category_name"],
                    "unit": row["unit"],
                    "base_price": row["base_price"],
                    "volatility": row["volatility"],
                    "item_quantity_range": row["item_quantity_range"],
                    "is_deleted": row["is_deleted"],
                    "updated_at": now,
                }
                await db["products"].update_one(
                    {"_id": existing["_id"]},
                    {"$set": encode_for_mongo(update_doc)},
                )
                updated += 1
            else:
                doc = {
                    "name": row["name"],
                    "category_id": row["category_id"],
                    "category_name": row["category_name"],
                    "unit": row["unit"],
                    "base_price": row["base_price"],
                    "volatility": row["volatility"],
                    "item_quantity_range": row["item_quantity_range"],
                    "is_deleted": row["is_deleted"],
                    "created_at": now,
                    "updated_at": now,
                }
                await db["products"].insert_one(encode_for_mongo(doc))
                created += 1
        if deactivate_candidates:
            result = await db["products"].update_many(
                {"name": {"$in": deactivate_candidates}},
                {"$set": {"is_deleted": True, "updated_at": now}},
            )
            deactivated = result.modified_count or 0
    else:
        for row in prepared:
            if row["name"] in existing_map:
                updated += 1
            else:
                created += 1

    return {
        "total": len(raw_rows),
        "valid": len(prepared),
        "skipped": max(len(raw_rows) - len(prepared), 0),
        "created": created,
        "updated": updated,
        "deactivated": deactivated,
        "warnings": warnings,
        "errors": [],
        "deactivate_candidates": deactivate_candidates,
        "applied": not dry_run,
        "dry_run": dry_run,
    }
