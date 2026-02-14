"""采购计划导出接口与单模板/导出设置配置。"""
from datetime import date, datetime
from zoneinfo import ZoneInfo
from decimal import Decimal, ROUND_HALF_UP
import io
import zipfile
from typing import Any

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from urllib.parse import quote

from app.db.mongo import get_database
from app.db.serializers import encode_for_mongo
from app.schemas.export_template import ExportTemplateUpdate
from app.core.response import ok

router = APIRouter(prefix="/api/procurement/exports", tags=["procurement-exports"])


def _month_range(start_year: int, start_month: int, end_year: int, end_month: int) -> list[tuple[int, int]]:
    """生成起止年月范围内的所有年月列表。"""
    months: list[tuple[int, int]] = []
    current = date(start_year, start_month, 1)
    end = date(end_year, end_month, 1)
    while current <= end:
        months.append((current.year, current.month))
        if current.month == 12:
            current = date(current.year + 1, 1, 1)
        else:
            current = date(current.year, current.month + 1, 1)
    return months


def _round_decimal(value: Decimal, precision: int) -> Decimal:
    """按精度对 Decimal 四舍五入。"""
    quant = Decimal("1") if precision <= 0 else Decimal("1").scaleb(-precision)
    return value.quantize(quant, rounding=ROUND_HALF_UP)


def _round_money(value: Decimal, precision: int) -> Decimal:
    """
    按精度显示金额，并做“最小显示单位保底（向上取整）”：
    - precision=0：小于 1 元按 1 元显示
    - precision=1：小于 0.1 元按 0.1 元显示
    - precision=2：小于 0.01 元按 0.01 元显示
    """
    if value == 0:
        return Decimal("0")
    min_unit = Decimal("1") if precision <= 0 else Decimal("1").scaleb(-precision)
    if 0 < abs(value) < min_unit:
        return min_unit if value > 0 else -min_unit
    return _round_decimal(value, precision)


def _coerce_decimal(value: Decimal | float | int | str | None) -> Decimal:
    """将任意数值转换为 Decimal，None 视为 0。"""
    if value is None:
        return Decimal("0")
    return Decimal(str(value))


FIELD_MAP = {
    "序号": "index",
    "时间": "date_text",
    "日期": "date_text",
    "物资及金额": "items_text",
    "小计": "day_total",
    "小计（元）": "day_total",
    "经手人": "handler",
    "证明人": "witness",
}


def _normalize_field(field: str) -> str:
    """将模板字段名映射为标准字段路径。"""
    cleaned = field.strip()
    return FIELD_MAP.get(cleaned, cleaned)


def _estimate_row_height(text: str, chars_per_line: int = 28) -> float:
    """根据单元格文本估算行高，避免显示被截断。"""
    if not text:
        return 16.0
    lines = (len(text) + chars_per_line - 1) // chars_per_line
    return max(16.0, 14.0 * lines)


def _default_template() -> dict[str, Any]:
    """生成单模板配置。"""
    return {
        "title": "{year}年{month:02d}月采购开支明细表",
        "columns": [
            {"label": "序号", "field": "index"},
            {"label": "时间", "field": "date_text"},
            {"label": "物资及金额", "field": "items_text"},
            {"label": "小计（元）", "field": "day_total"},
            {"label": "经手人", "field": "handler"},
            {"label": "证明人", "field": "witness"},
        ],
    }


def _serialize_template(doc: dict[str, Any]) -> dict[str, Any]:
    """序列化模板文档为接口输出结构。"""
    result = dict(doc)
    result["id"] = str(result.pop("_id"))
    return result


def _get_field_value(row: dict[str, Any], path: str) -> Any:
    """按字段路径读取行数据，缺失则返回空字符串。"""
    if not path:
        return ""
    path = _normalize_field(path)
    if path in row:
        return row[path]
    value: Any = row
    for part in path.split("."):
        if isinstance(value, dict) and part in value:
            value = value[part]
        else:
            return ""
    return value


def _format_title(template: dict[str, Any], year: int, month: int) -> str:
    """根据模板与年月格式化表头标题。"""
    title = template.get("title") or ""
    if not title:
        return f"{year}年{month:02d}月采购开支明细表"
    return title.format(year=year, month=month)

def _format_plan_date(value: str | None) -> str:
    """格式化计划日期为 MM月DD日。"""
    if not value:
        return ""
    return date.fromisoformat(value).strftime("%m月%d日")

def _build_items_text_and_day_total(plan: dict[str, Any], precision: int) -> tuple[str, Decimal]:
    """构建物资及金额文本，并计算单日小计（按导出精度）。"""
    items_text: list[str] = []
    item_price_precision = max(precision, 2)
    rounded_item_amounts: list[Decimal] = []
    for item in plan.get("items", []):
        name = item.get("name", "")
        amount = _coerce_decimal(item.get("amount"))
        if amount == 0:
            price = _round_decimal(
                _coerce_decimal(item.get("price")),
                item_price_precision,
            )
            items_text.append(f"{name}{format(price, f'.{item_price_precision}f')}元")
            continue
        display_amount = _round_money(amount, precision)
        rounded_item_amounts.append(display_amount)
        items_text.append(f"{name}{format(display_amount, f'.{precision}f')}元")

    if rounded_item_amounts:
        day_total = sum(rounded_item_amounts, Decimal("0"))
    else:
        day_total = _round_money(_coerce_decimal(plan.get("total_amount")), precision)

    return "、".join(items_text), day_total


def _resolve_columns(template: dict[str, Any]) -> list[dict[str, Any]]:
    """解析模板列配置，缺省时返回默认列。"""
    columns = template.get("columns") or []
    if not columns:
        columns = _default_template()["columns"]
    return columns


def _column_width(label: str, field: str) -> int:
    """根据列类型与字段决定列宽。"""
    field = _normalize_field(field)
    if "物资" in label or field == "items_text":
        return 64
    if "时间" in label or field in {"date_text", "date"}:
        return 12
    if "序号" in label or field == "index":
        return 6
    if "小计" in label or field in {"day_total", "total_amount"}:
        return 12
    return 10


def _money_number_format(precision: int) -> str:
    """按精度生成金额显示格式。"""
    if precision <= 0:
        return "0"
    return "0." + ("0" * precision)

def _format_money_display(value: Decimal, precision: int) -> str:
    """按精度格式化金额显示字符串。"""
    rounded = _round_money(value, precision)
    if precision <= 0:
        return format(rounded, ".0f")
    return format(rounded, f".{precision}f")


def _build_workbook(
    year: int,
    month: int,
    plans: list[dict[str, Any]],
    precision: int,
    template: dict[str, Any],
) -> Workbook:
    """构建单月采购清单的 Excel 工作簿。"""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = f"{year}-{month:02d}"

    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    wrap_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    columns = _resolve_columns(template)
    title = _format_title(template, year, month)

    ws.append([title])
    ws.append([])
    ws.append([col.get("label", "") for col in columns])

    last_col = get_column_letter(len(columns))
    ws.merge_cells(f"A1:{last_col}2")
    ws["A1"].alignment = center_align

    ws["A1"].font = title_font
    for col_idx in range(1, len(columns) + 1):
        ws.cell(row=3, column=col_idx).font = header_font

    for idx, col in enumerate(columns, start=1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = _column_width(col.get("label", ""), col.get("field", ""))

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 0
    ws.row_dimensions[3].height = 18

    month_total = Decimal("0")

    for idx, plan in enumerate(plans, start=1):
        items_text, day_total_display = _build_items_text_and_day_total(plan, precision)
        month_total += day_total_display

        row_data: dict[str, Any] = {
            "index": idx,
            "date_text": _format_plan_date(plan.get("date")),
            "items_text": items_text,
            "day_total": day_total_display,
            "handler": "",
            "witness": "",
            **plan,
        }

        row = []
        for col in columns:
            field = col.get("field", "")
            value = _get_field_value(row_data, field)
            if isinstance(value, Decimal):
                value = _round_decimal(value, precision)
            row.append(value)

        ws.append(row)
        total_format = _money_number_format(precision)
        for col_idx, col in enumerate(columns, start=1):
            label = col.get("label", "")
            field = col.get("field", "")
            if "小计" in label or field in {"day_total", "total_amount"}:
                ws.cell(row=ws.max_row, column=col_idx).number_format = total_format
        row_height_text = next(
            (
                _get_field_value(row_data, col.get("field", ""))
                for col in columns
                if col.get("field", "") in {"items_text"}
                or "物资" in col.get("label", "")
            ),
            "",
        )
        ws.row_dimensions[ws.max_row].height = _estimate_row_height(str(row_height_text))

    total_row = []
    for col in columns:
        label = col.get("label", "")
        if "序号" in label:
            total_row.append("总计")
        elif "小计" in label or col.get("field") in {"day_total", "total_amount"}:
            total_row.append(_round_money(month_total, precision))
        else:
            total_row.append("")

    ws.append(total_row)
    total_format = _money_number_format(precision)
    for col_idx, col in enumerate(columns, start=1):
        label = col.get("label", "")
        field = col.get("field", "")
        if "小计" in label or field in {"day_total", "total_amount"}:
            ws.cell(row=ws.max_row, column=col_idx).number_format = total_format
    ws.row_dimensions[ws.max_row].height = 18

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)


    left_columns = {
        idx + 1
        for idx, col in enumerate(columns)
        if col.get("field", "") in {"items_text"} or "物资" in col.get("label", "")
    }
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            if cell.column in left_columns:
                cell.alignment = wrap_left
            else:
                cell.alignment = center_align
            cell.border = border

    if ws.max_row >= 4:
        ws.cell(row=ws.max_row, column=1).font = header_font
        for idx, col in enumerate(columns, start=1):
            if "小计" in col.get("label", "") or col.get("field") in {"day_total", "total_amount"}:
                ws.cell(row=ws.max_row, column=idx).font = header_font
    return wb

def _build_preview_rows(
    plans: list[dict[str, Any]],
    precision: int,
    max_rows: int,
) -> tuple[list[dict[str, Any]], Decimal]:
    """构建导出模板预览行与月度合计。"""
    rows: list[dict[str, Any]] = []
    month_total = Decimal("0")
    limited_plans = plans[:max_rows] if max_rows > 0 else plans

    for idx, plan in enumerate(limited_plans, start=1):
        items_text, day_total = _build_items_text_and_day_total(plan, precision)
        month_total += day_total

        rows.append(
            {
                "index": idx,
                "date_text": _format_plan_date(plan.get("date")),
                "items_text": items_text,
                "day_total": _format_money_display(day_total, precision),
                "handler": "",
                "witness": "",
            }
        )

    return rows, month_total


async def _resolve_single_template() -> dict[str, Any]:
    """获取单模板配置，不存在则返回默认值。"""
    db = get_database()
    doc = await db["export_templates"].find_one({})
    if not doc:
        return _default_template()
    return doc


@router.post("")
async def export_zip(
    start_year: int,
    start_month: int,
    end_year: int,
    end_month: int,
) -> StreamingResponse:
    """按年月范围导出采购清单，返回 ZIP 文件流。"""
    db = get_database()
    settings = await db["settings"].find_one({"key": "global"})
    precision = int((settings or {}).get("export_precision", 2))

    time_tag = datetime.now(ZoneInfo("Asia/Shanghai")).strftime("%Y%m%d_%H%M%S")

    months = _month_range(start_year, start_month, end_year, end_month)
    template = await _resolve_single_template()
    workbooks: list[tuple[str, bytes]] = []

    for year, month in months:
        year_month = f"{year}-{month:02d}"
        cursor = db["procurement_plans"].find({"year_month": year_month}).sort("date", 1)
        plans: list[dict[str, Any]] = [doc async for doc in cursor]

        if not plans:
            continue

        wb = _build_workbook(
            year,
            month,
            plans,
            precision,
            template,
        )
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        filename = f"{year}年{month:02d}月采购清单.xlsx"
        workbooks.append((filename, excel_buffer.read()))

    if not workbooks:
        raise HTTPException(status_code=409, detail="当前选中时间区间无采购计划数据")

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for filename, payload in workbooks:
            zip_file.writestr(filename, payload)

    zip_buffer.seek(0)
    zip_name = f"采购清单_{start_year}{start_month:02d}_{end_year}{end_month:02d}_{time_tag}.zip"
    ascii_name = f"procurement_{start_year}{start_month:02d}_{end_year}{end_month:02d}_{time_tag}.zip"
    encoded_name = quote(zip_name)
    disposition = f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{encoded_name}"
    headers = {"Content-Disposition": disposition}
    return StreamingResponse(zip_buffer, media_type="application/zip", headers=headers)


@router.get("/settings")
async def get_export_settings() -> dict:
    """获取导出相关设置。"""
    db = get_database()
    doc = await db["settings"].find_one({"key": "global"})
    precision = int((doc or {}).get("export_precision", 2))
    return ok({"export_precision": precision})


@router.put("/settings")
async def update_export_settings(payload: dict[str, Any]) -> dict:
    """更新导出相关设置（当前仅导出金额精度）。"""
    raw_precision = payload.get("export_precision")
    if raw_precision is None:
        raise HTTPException(status_code=400, detail="缺少导出金额精度")
    try:
        export_precision = int(raw_precision)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="导出金额精度仅支持 0、1、2")
    if export_precision not in {0, 1, 2}:
        raise HTTPException(status_code=400, detail="导出金额精度仅支持 0、1、2")

    db = get_database()
    now = datetime.utcnow()
    await db["settings"].update_one(
        {"key": "global"},
        {
            "$set": {"key": "global", "export_precision": export_precision, "updated_at": now},
            "$setOnInsert": {"created_at": now},
        },
        upsert=True,
    )
    return ok({"status": "成功"})


@router.get("/templates/single")
async def get_single_template():
    """获取单模板配置，不存在则创建默认模板。"""
    db = get_database()
    doc = await db["export_templates"].find_one({})
    if not doc:
        now = datetime.utcnow()
        template = _default_template()
        template["created_at"] = now
        template["updated_at"] = now
        result = await db["export_templates"].insert_one(encode_for_mongo(template))
        template["_id"] = result.inserted_id
        doc = template
    return ok(_serialize_template(doc))

@router.put("/templates/single")
async def update_single_template(payload: ExportTemplateUpdate) -> dict:
    """更新单模板配置。"""
    db = get_database()
    doc = await db["export_templates"].find_one({})
    if not doc:
        raise HTTPException(status_code=404, detail="模板不存在")
    update = {k: v for k, v in payload.model_dump().items() if v is not None and k in {"title", "columns"}}
    update = encode_for_mongo(update)
    update["updated_at"] = datetime.utcnow()
    await db["export_templates"].update_one({"_id": doc["_id"]}, {"$set": update})
    return {"status": "成功", "id": str(doc["_id"])}

@router.get("/preview")
async def preview_export(
    year: int,
    month: int,
    max_rows: int = 3,
) -> dict:
    """预览导出模板数据，金额与小计按导出精度计算。"""
    db = get_database()
    settings = await db["settings"].find_one({"key": "global"})
    precision = int((settings or {}).get("export_precision", 2))
    precision = precision if precision in {0, 1, 2} else 2

    year_month = f"{year}-{month:02d}"
    cursor = db["procurement_plans"].find({"year_month": year_month}).sort("date", 1)
    plans: list[dict[str, Any]] = [doc async for doc in cursor]

    rows, month_total = _build_preview_rows(plans, precision, max_rows)
    return ok(
        {
            "precision": precision,
            "rows": rows,
            "month_total": _format_money_display(month_total, precision),
        }
    )
